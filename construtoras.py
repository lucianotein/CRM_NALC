import os
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from pathlib import Path
from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from django.db import transaction

from accounts.models import Account
from projects.models import Project


# =========================
# CONFIG
# =========================
ARQUIVO_EXCEL = Path("Construtoras.xlsx")  # coloque o arquivo na raiz do projeto
USERNAME_DONO = "aristeu"
NOME_SHEET = None  # None = usa a primeira aba


# =========================
# HELPERS
# =========================
def limpar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def normalizar_chave(valor):
    return limpar_texto(valor).upper()


# =========================
# IMPORTAÇÃO
# =========================
@transaction.atomic
def run():
    if not ARQUIVO_EXCEL.exists():
        raise FileNotFoundError(
            f"Arquivo não encontrado: {ARQUIVO_EXCEL.resolve()}"
        )

    User = get_user_model()

    try:
        owner = User.objects.get(username=USERNAME_DONO)
    except User.DoesNotExist:
        raise Exception(f"Usuário '{USERNAME_DONO}' não encontrado.")

    wb = load_workbook(ARQUIVO_EXCEL, data_only=True)

    if NOME_SHEET:
        if NOME_SHEET not in wb.sheetnames:
            raise Exception(
                f"Planilha '{NOME_SHEET}' não encontrada. Disponíveis: {wb.sheetnames}"
            )
        ws = wb[NOME_SHEET]
    else:
        ws = wb[wb.sheetnames[0]]

    contas_criadas = 0
    contas_existentes = 0
    empreendimentos_criados = 0
    empreendimentos_existentes = 0
    linhas_ignoradas = 0

    contas_cache = {}

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        construtora_raw = row[0] if len(row) > 0 else None
        empreendimento_raw = row[1] if len(row) > 1 else None

        construtora = limpar_texto(construtora_raw)
        empreendimento = limpar_texto(empreendimento_raw)

        # Ignora cabeçalho comum
        if idx == 1 and normalizar_chave(construtora) in {"CONSTRUTORA", "CONSTRUTORA"}:
            continue

        # Ignora linhas sem construtora
        if not construtora:
            linhas_ignoradas += 1
            continue

        conta_key = normalizar_chave(construtora)

        if conta_key in contas_cache:
            conta = contas_cache[conta_key]
        else:
            conta, created = Account.objects.get_or_create(
                owner=owner,
                name=construtora,
                defaults={
                    "cnpj": "",
                    "city": "",
                    "state": "",
                    "is_active": True,
                },
            )

            if created:
                contas_criadas += 1
                print(f"[CONTA CRIADA] {conta.name}")
            else:
                contas_existentes += 1
                print(f"[CONTA EXISTENTE] {conta.name}")

            contas_cache[conta_key] = conta

        # Se não houver empreendimento, cria só a construtora
        if not empreendimento:
            continue

        projeto, created = Project.objects.get_or_create(
            account=conta,
            name=empreendimento,
            defaults={
                "city": "",
                "state": "",
                "obra_entrega_prevista": "",
                "notes": "",
            },
        )

        if created:
            empreendimentos_criados += 1
            print(f"   -> [EMPREENDIMENTO CRIADO] {projeto.name}")
        else:
            empreendimentos_existentes += 1
            print(f"   -> [EMPREENDIMENTO EXISTENTE] {projeto.name}")

    print("\n=========================")
    print("IMPORTAÇÃO FINALIZADA")
    print("=========================")
    print(f"Contas criadas: {contas_criadas}")
    print(f"Contas existentes: {contas_existentes}")
    print(f"Empreendimentos criados: {empreendimentos_criados}")
    print(f"Empreendimentos existentes: {empreendimentos_existentes}")
    print(f"Linhas ignoradas: {linhas_ignoradas}")


if __name__ == "__main__":
    run()
